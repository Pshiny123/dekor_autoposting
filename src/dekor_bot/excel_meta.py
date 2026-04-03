from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime, timezone
import json
import os
from pathlib import Path
from urllib.parse import urlparse

import pandas as pd
from openpyxl import load_workbook
try:
    import gspread
except Exception:  # pragma: no cover
    gspread = None  # type: ignore[assignment]


def _norm(s: str) -> str:
    return str(s).strip().lstrip("\ufeff").strip().casefold()


def _kv_key_loose(s: str) -> str:
    """Ключ без регистра, BOM, пробелов и подчёркиваний — чтобы сработало и «Post index»."""
    return _norm(s).replace(" ", "").replace("_", "")


def _kv_lookup_ci(kv: dict[str, str], *keyword_norms: str) -> str:
    """
    Ищем значение по Key/Value листу без учёта регистра и пробелов в имени ключа.
    """
    norm_to_val: dict[str, str] = {}
    for k, v in kv.items():
        nk = _kv_key_loose(k)
        if nk:
            norm_to_val[nk] = "" if v is None else str(v).strip()
    for want in keyword_norms:
        w = _kv_key_loose(want)
        if w in norm_to_val:
            return norm_to_val[w]
    return ""


def _is_google_sheets_url(source: str | Path) -> bool:
    return str(source).strip().lower().startswith("https://docs.google.com/spreadsheets/")


def _extract_gsheet_id(url: str) -> str:
    parsed = urlparse(url)
    parts = [p for p in parsed.path.split("/") if p]
    try:
        i = parts.index("d")
        return parts[i + 1]
    except Exception as exc:
        raise ValueError("Не удалось извлечь spreadsheet id из Google Sheets URL.") from exc


def _get_gspread_client():
    if gspread is None:
        raise RuntimeError("Для Google Sheets установите зависимости: gspread и google-auth.")
    raw_inline = os.getenv("GOOGLE_SERVICE_ACCOUNT_JSON_INLINE", "")
    inline_json = raw_inline.strip().lstrip("\ufeff")
    if inline_json:
        try:
            info = json.loads(inline_json)
            return gspread.service_account_from_dict(info)
        except json.JSONDecodeError as exc:
            hint = (
                "Частые причины: (1) JSON в .env разбит на несколько строк — должна быть РОВНО одна строка, "
                "в private_key только \\n как два символа; (2) в строке есть неэкранированные кавычки; "
                "(3) после значения на той же строке идёт # комментарий — обрежет JSON. "
                "Надёжнее: сохраните ключ в .json файл и задайте GOOGLE_SERVICE_ACCOUNT_JSON=/полный/путь/к/ключу.json"
            )
            raise ValueError(
                f"GOOGLE_SERVICE_ACCOUNT_JSON_INLINE: ошибка JSON — {exc.msg} (позиция {exc.pos}). {hint}"
            ) from exc
        except Exception as exc:
            raise ValueError(
                "GOOGLE_SERVICE_ACCOUNT_JSON_INLINE: не удалось разобрать ключ. Проверьте однострочный JSON или используйте файл GOOGLE_SERVICE_ACCOUNT_JSON=..."
            ) from exc
    creds_path = os.getenv("GOOGLE_SERVICE_ACCOUNT_JSON", "").strip()
    if creds_path:
        p = Path(creds_path).expanduser().resolve()
        if not p.exists():
            raise FileNotFoundError(f"Файл service account не найден: {p}")
        return gspread.service_account(filename=str(p))
    raise RuntimeError("Не задан GOOGLE_SERVICE_ACCOUNT_JSON или GOOGLE_SERVICE_ACCOUNT_JSON_INLINE.")


def _open_gsheet(source: str | Path):
    sid = _extract_gsheet_id(str(source))
    return _get_gspread_client().open_by_key(sid)


def _gsheet_worksheet_by_title(spreadsheet, title: str):
    """Имя листа в UI может отличаться регистром — gspread ищет только точное совпадение."""
    want = _norm(title)
    for ws in spreadsheet.worksheets():
        if _norm(ws.title) == want:
            return ws
    names = [ws.title for ws in spreadsheet.worksheets()]
    raise ValueError(f"Лист «{title}» не найден в таблице. Доступные листы: {names}")


def _worksheet_to_df(ws) -> pd.DataFrame:
    values = ws.get_all_values()
    if not values:
        return pd.DataFrame()
    header = values[0]
    rows = values[1:] if len(values) > 1 else []
    return pd.DataFrame(rows, columns=header)


def _sheet_exists(xlsx_path: str | Path, sheet_name: str) -> bool:
    if _is_google_sheets_url(xlsx_path):
        try:
            sh = _open_gsheet(xlsx_path)
            _gsheet_worksheet_by_title(sh, sheet_name)
            return True
        except Exception:
            return False
    xlsx_path = Path(xlsx_path)
    xls = pd.ExcelFile(xlsx_path)
    return any(_norm(n) == _norm(sheet_name) for n in xls.sheet_names)


def _pick_sheet_name(xlsx_path: str | Path, desired: str) -> str:
    xlsx_path = Path(xlsx_path)
    xls = pd.ExcelFile(xlsx_path)
    desired_norm = _norm(desired)
    for name in xls.sheet_names:
        if _norm(name) == desired_norm:
            return name
    raise ValueError(f"Лист '{desired}' не найден. Доступные листы: {xls.sheet_names}")


def _read_kv_sheet(xlsx_path: str | Path, sheet_name: str) -> dict[str, str]:
    if _is_google_sheets_url(xlsx_path):
        sh = _open_gsheet(xlsx_path)
        ws = _gsheet_worksheet_by_title(sh, sheet_name)
        df = _worksheet_to_df(ws)
        sheet = sheet_name
    else:
        sheet = _pick_sheet_name(xlsx_path, sheet_name)
        df = pd.read_excel(Path(xlsx_path), sheet_name=sheet)
    if "Key" not in df.columns or "Value" not in df.columns:
        raise ValueError(f"Лист '{sheet}' должен содержать колонки Key и Value.")
    out: dict[str, str] = {}
    for _, row in df.iterrows():
        k = row.get("Key")
        v = row.get("Value")
        if k is None or pd.isna(k):
            continue
        key = str(k).strip().lstrip("\ufeff").strip()
        if not key:
            continue
        if v is None or pd.isna(v):
            out[key] = ""
        else:
            out[key] = str(v)
    return out


def _utc_now() -> datetime:
    return datetime.now(timezone.utc)


def _dt_to_iso(dt: datetime) -> str:
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    return dt.astimezone(timezone.utc).isoformat()


def _dt_from_iso(s: str) -> datetime:
    dt = datetime.fromisoformat(s)
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    return dt.astimezone(timezone.utc)


def _parse_post_index_value(raw: str) -> int:
    """
    Google Sheets / pandas часто дают число как строку '34.0'.
    int('34.0') падает — из-за этого раньше индекс сбрасывался на 1 и постился только первый пост.
    """
    s = (raw or "").strip()
    if not s:
        return 1
    try:
        x = float(s.replace(",", "."))
        if x < 1:
            return 1
        return max(1, int(x))
    except ValueError:
        return 1


def _norm_queue_post_id(v: object) -> str | None:
    """PostID из таблицы: 5 и 5.0 -> строка '5' для совпадения с колонкой ID в Posts."""
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return None
    if isinstance(v, bool):
        return None
    if isinstance(v, int):
        return str(v)
    if isinstance(v, float):
        if v == int(v):
            return str(int(v))
        return str(v).strip()
    s = str(v).strip()
    if not s:
        return None
    try:
        xf = float(s.replace(",", "."))
        if xf == int(xf):
            return str(int(xf))
    except ValueError:
        pass
    return s


@dataclass(frozen=True)
class ExcelState:
    # 1-based индекс шага в Queue
    post_index: int
    last_posted_at: datetime | None


def has_meta_sheets(xlsx_path: str | Path) -> bool:
    return _sheet_exists(xlsx_path, "State") and _sheet_exists(xlsx_path, "Queue") and _sheet_exists(xlsx_path, "Settings")


def read_settings_chat_id(xlsx_path: str | Path) -> str:
    kv = _read_kv_sheet(xlsx_path, "Settings")
    chat_id = kv.get("chat_id", "").strip()
    if not chat_id:
        raise ValueError("В листе Settings не найден ключ chat_id.")
    return chat_id


def read_frequency_days(xlsx_path: str | Path) -> int | None:
    """
    Поддерживаем 2 формата:
    1) Лист `Frequency` с колонкой `frequency` и одним значением (1/2/3)
    2) Лист `Frequency` в формате Key/Value, где Key=frequency
    """
    if not _sheet_exists(xlsx_path, "Frequency"):
        return None
    if _is_google_sheets_url(xlsx_path):
        sh = _open_gsheet(xlsx_path)
        df = _worksheet_to_df(_gsheet_worksheet_by_title(sh, "Frequency"))
    else:
        sheet = _pick_sheet_name(xlsx_path, "Frequency")
        df = pd.read_excel(Path(xlsx_path), sheet_name=sheet)

    # Частый случай: в листе одна строка вида "Frequency value | 2",
    # и pandas считает её заголовком, оставляя 0 строк данных.
    # Тогда значение можно достать из имени второй колонки.
    if df.shape[0] == 0 and len(df.columns) >= 2:
        c0 = str(df.columns[0]).strip().casefold()
        c1 = str(df.columns[1]).strip()
        if "frequency" in c0:
            try:
                n = int(c1)
            except Exception:
                n = None
            if n is not None:
                if n < 1:
                    raise ValueError("Frequency: значение frequency должно быть >= 1.")
                return n

    # key/value формат
    if "Key" in df.columns and "Value" in df.columns:
        kv = _read_kv_sheet(xlsx_path, "Frequency")
        raw = (kv.get("frequency") or kv.get("Frequency") or "").strip()
        if not raw:
            return None
        try:
            n = int(raw)
        except Exception:
            raise ValueError("Frequency: значение frequency должно быть числом 1/2/3.")
        if n < 1:
            raise ValueError("Frequency: значение frequency должно быть >= 1.")
        return n

    # одна колонка frequency
    if "frequency" in df.columns:
        if df.shape[0] < 1:
            return None
        raw = df.iloc[0]["frequency"]
        if raw is None or pd.isna(raw):
            return None
        try:
            n = int(str(raw).strip())
        except Exception:
            raise ValueError("Frequency: значение frequency должно быть числом 1/2/3.")
        if n < 1:
            raise ValueError("Frequency: значение frequency должно быть >= 1.")
        return n

    # если лист есть, но формат другой — явно скажем
    raise ValueError("Лист Frequency должен иметь колонку 'frequency' или формат Key/Value с Key=frequency.")


def read_queue_post_ids(xlsx_path: str | Path) -> list[str]:
    if _is_google_sheets_url(xlsx_path):
        sh = _open_gsheet(xlsx_path)
        df = _worksheet_to_df(_gsheet_worksheet_by_title(sh, "Queue"))
    else:
        sheet = _pick_sheet_name(xlsx_path, "Queue")
        df = pd.read_excel(Path(xlsx_path), sheet_name=sheet)
    if "PostID" not in df.columns:
        raise ValueError("Лист Queue должен содержать колонку PostID.")
    post_ids: list[str] = []
    for _, row in df.iterrows():
        v = row.get("PostID")
        s = _norm_queue_post_id(v)
        if s:
            post_ids.append(s)
    if not post_ids:
        raise ValueError("Queue пустой: не найдено ни одного PostID.")
    return post_ids


def read_state(xlsx_path: str | Path) -> ExcelState:
    kv = _read_kv_sheet(xlsx_path, "State")
    raw = _kv_lookup_ci(kv, "Postindex", "PostIndex", "post_index", "step", "Step")
    post_index = _parse_post_index_value(str(raw))

    last_raw = _kv_lookup_ci(kv, "LastPostedAt", "LastPosted", "last_posted_at").strip()
    last_dt = _dt_from_iso(last_raw) if last_raw else None
    return ExcelState(post_index=post_index, last_posted_at=last_dt)


def write_state(xlsx_path: str | Path, *, post_index: int, last_posted_at: datetime | None) -> None:
    if _is_google_sheets_url(xlsx_path):
        sh = _open_gsheet(xlsx_path)
        ws = _gsheet_worksheet_by_title(sh, "State")
        rows = ws.get_all_values()
        if not rows:
            ws.update("A1:B1", [["Key", "Value"]])
            rows = ws.get_all_values()
        header = rows[0]
        key_col = next((i + 1 for i, h in enumerate(header) if _norm(str(h)) == "key"), None)
        value_col = next((i + 1 for i, h in enumerate(header) if _norm(str(h)) == "value"), None)
        if key_col is None or value_col is None:
            raise ValueError("Лист State должен иметь заголовки Key и Value в первой строке.")

        def row_key_matches(cell_val: str, key: str) -> bool:
            a, b = str(cell_val), str(key)
            return _norm(a) == _norm(b) or _kv_key_loose(a) == _kv_key_loose(b)

        def upsert(key: str, value: str) -> None:
            cur_rows = ws.get_all_values()
            for r_idx, row in enumerate(cur_rows[1:], start=2):
                cur = row[key_col - 1] if len(row) >= key_col else ""
                if row_key_matches(cur, key):
                    ws.update_cell(r_idx, value_col, value)
                    return
            width = max(len(cur_rows[0]), value_col)
            new_row = [""] * width
            new_row[key_col - 1] = key
            new_row[value_col - 1] = value
            ws.append_row(new_row, value_input_option="RAW")

        upsert("Postindex", str(int(post_index)))
        upsert("LastPostedAt", _dt_to_iso(last_posted_at or _utc_now()))
        return

    xlsx_path = Path(xlsx_path)
    sheet = _pick_sheet_name(xlsx_path, "State")
    wb = load_workbook(xlsx_path)
    ws = wb[sheet]

    # Ожидаем заголовки Key/Value в первой строке.
    # Ищем строку с Postindex и LastPostedAt; если нет — добавляем.
    key_col = None
    value_col = None
    for col in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=col).value
        if isinstance(v, str) and _norm(v) == "key":
            key_col = col
        if isinstance(v, str) and _norm(v) == "value":
            value_col = col
    if key_col is None or value_col is None:
        raise ValueError("Лист State должен иметь заголовки Key и Value в первой строке.")

    def row_key_matches_cell(k: object, key: str) -> bool:
        if k is None:
            return False
        a = str(k).strip()
        return _norm(a) == _norm(key) or _kv_key_loose(a) == _kv_key_loose(key)

    def upsert(key: str, value: str) -> None:
        for r in range(2, ws.max_row + 1):
            k = ws.cell(row=r, column=key_col).value
            if isinstance(k, str) and row_key_matches_cell(k, key):
                ws.cell(row=r, column=value_col).value = value
                return
            if k is not None and not isinstance(k, str) and str(k).strip() and row_key_matches_cell(k, key):
                ws.cell(row=r, column=value_col).value = value
                return
        r = ws.max_row + 1
        ws.cell(row=r, column=key_col).value = key
        ws.cell(row=r, column=value_col).value = value

    upsert("Postindex", str(int(post_index)))
    upsert("LastPostedAt", _dt_to_iso(last_posted_at or _utc_now()))
    wb.save(xlsx_path)

