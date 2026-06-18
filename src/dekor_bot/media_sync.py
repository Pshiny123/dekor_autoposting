"""
Автосинхронизация медиа: Google Sheets / Excel ↔ Yandex Object Storage.

При добавлении, изменении или удалении ссылки в Photo/Video:
  - внешняя ссылка → скачать, залить в posts/{ID}/{колонка}.ext, подставить URL бакета в таблицу;
  - пустая ячейка → удалить файл(ы) из бакета;
  - смена ссылки → удалить старый объект, залить новый.
"""
from __future__ import annotations

import json
import logging
import os
import tempfile
from dataclasses import dataclass
from pathlib import Path
from threading import Lock

from openpyxl import load_workbook

from .excel_meta import _extract_gsheet_id, _get_gspread_client, _gsheet_worksheet_by_title, _is_google_sheets_url
from .excel_posts import _pick_sheet_name, _read_gsheet_df
from .video_compress import compress_enabled, max_video_bytes, prepare_video_for_telegram
from .yandex_storage import (
    bucket_name,
    delete_cell,
    delete_keys,
    download_media,
    download_object,
    get_s3_client,
    guess_ext,
    is_http_url,
    is_yandex_url,
    list_objects,
    object_key,
    object_size,
    upload_file,
    yandex_configured,
    yandex_public_url,
    yandex_url_to_key,
)

logger = logging.getLogger(__name__)

_state_lock = Lock()
MEDIA_COLS = [f"Photo{i}" for i in range(1, 11)] + [f"Video{i}" for i in range(1, 11)]


@dataclass(frozen=True)
class MediaCell:
    post_id: str
    col: str
    url: str
    kind: str  # photo | video


def _env_bool(name: str, default: bool) -> bool:
    v = os.getenv(name)
    if v is None:
        return default
    return v.strip().lower() in {"1", "true", "yes", "y", "on"}


def auto_sync_enabled() -> bool:
    if not yandex_configured():
        return False
    return _env_bool("YANDEX_AUTO_SYNC", True)


def _state_path() -> Path:
    data_dir = Path((os.getenv("DATA_DIR") or "data").strip())
    raw = (os.getenv("MEDIA_SYNC_STATE_PATH") or str(data_dir / "media_sync_state.json")).strip()
    return Path(raw)


def _load_state() -> dict:
    path = _state_path()
    if not path.is_file():
        return {"cells": {}}
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
        if isinstance(data, dict) and isinstance(data.get("cells"), dict):
            return data
    except (OSError, json.JSONDecodeError):
        pass
    return {"cells": {}}


def _save_state(state: dict) -> None:
    path = _state_path()
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp = path.with_suffix(path.suffix + ".tmp")
    tmp.write_text(json.dumps(state, ensure_ascii=False, indent=2), encoding="utf-8")
    tmp.replace(path)


def _cell_key(post_id: str, col: str) -> str:
    return f"{post_id}:{col}"


def _is_video_col(col: str) -> bool:
    return col.startswith("Video")


def _prepare_local_media(local: Path, kind: str) -> tuple[Path, str]:
    if kind == "video":
        prepared = prepare_video_for_telegram(local)
        ext = prepared.suffix.lstrip(".").lower() or "mp4"
        return prepared, ext
    return local, guess_ext("", kind, local_path=local)


def _recompress_bucket_video_if_needed(
    s3,
    bucket: str,
    post_id: str,
    col: str,
    ykey: str,
) -> bool:
    if not _is_video_col(col) or not compress_enabled():
        return False
    size = object_size(s3, bucket, ykey)
    if size is None or size <= max_video_bytes():
        return False

    with tempfile.TemporaryDirectory(prefix="dekor_recompress_") as tmp:
        src = Path(tmp) / "source.mp4"
        download_object(s3, bucket, ykey, src)
        prepared, ext = _prepare_local_media(src, "video")
        new_key = object_key(post_id, col, ext)
        if new_key != ykey:
            delete_keys(s3, bucket, [ykey])
        upload_file(s3, bucket, prepared, new_key)
        logger.info(
            "Media sync: #%s %s пересжато в бакете %.1f → %.1f МБ",
            post_id,
            col,
            size / (1024 * 1024),
            prepared.stat().st_size / (1024 * 1024),
        )
    return True


def load_media_cells(source: str, sheet_name: str) -> list[MediaCell]:
    source_s = str(source).strip()
    if _is_google_sheets_url(source_s):
        actual = _pick_sheet_name(source_s, sheet_name)
        df = _read_gsheet_df(source_s, actual)
        if df is None:
            raise ValueError("Не удалось прочитать лист Posts.")
    else:
        import pandas as pd

        actual = _pick_sheet_name(source_s, sheet_name)
        df = pd.read_excel(Path(source_s).resolve(), sheet_name=actual)

    if "ID" not in df.columns:
        raise ValueError("В листе Posts нужна колонка ID.")

    cells: list[MediaCell] = []
    for _, row in df.iterrows():
        raw_id = row.get("ID", "")
        if raw_id is None or str(raw_id).strip() == "" or str(raw_id).lower() == "nan":
            continue
        post_id = str(raw_id).strip().lstrip("\ufeff")
        for col in MEDIA_COLS:
            raw = row.get(col, "")
            if raw is None or str(raw).strip() == "" or str(raw).lower() == "nan":
                url = ""
            else:
                url = str(raw).strip()
            kind = "video" if col.startswith("Video") else "photo"
            cells.append(MediaCell(post_id=post_id, col=col, url=url, kind=kind))
    return cells


def _should_sync_url(url: str, bucket: str) -> bool:
    if not url:
        return False
    if is_yandex_url(url, bucket):
        return False
    if is_http_url(url):
        return True
    return Path(url).is_file() or (Path.cwd() / url).is_file()


def _parse_bucket_cell_key(key: str) -> tuple[str, str] | None:
    # posts/52/Video1.mp4
    if not key.startswith("posts/"):
        return None
    rest = key[len("posts/") :]
    parts = rest.split("/", 1)
    if len(parts) != 2:
        return None
    post_id, filename = parts
    col = filename.rsplit(".", 1)[0]
    if col in MEDIA_COLS:
        return post_id, col
    return None


def update_sheet_cells(
    source: str,
    sheet_name: str,
    updates: list[tuple[str, str, str]],
) -> None:
    """updates: (post_id, column_name, new_url)"""
    if not updates:
        return

    if _is_google_sheets_url(source):
        from gspread.utils import rowcol_to_a1

        client = _get_gspread_client()
        sh = client.open_by_key(_extract_gsheet_id(source))
        ws = _gsheet_worksheet_by_title(sh, _pick_sheet_name(source, sheet_name))
        values = ws.get_all_values()
        header = values[0]
        col_index = {name: idx for idx, name in enumerate(header)}
        id_col = col_index["ID"]
        by_post_row: dict[str, int] = {}
        for r_idx, row in enumerate(values[1:], start=2):
            if id_col < len(row):
                pid = str(row[id_col]).strip()
                if pid:
                    by_post_row[pid] = r_idx

        batch_data: list[dict] = []
        for post_id, col_name, new_url in updates:
            if col_name not in col_index:
                logger.warning("Media sync: колонка %s не найдена", col_name)
                continue
            row = by_post_row.get(str(post_id))
            if row is None:
                logger.warning("Media sync: пост #%s не найден для обновления %s", post_id, col_name)
                continue
            cell_a1 = rowcol_to_a1(row, col_index[col_name] + 1)
            batch_data.append({"range": cell_a1, "values": [[new_url]]})
            logger.info("Media sync: таблица #%s %s → %s", post_id, col_name, new_url)

        if batch_data:
            ws.batch_update(batch_data, value_input_option="RAW")
        return

    xlsx_path = Path(source).resolve()
    wb = load_workbook(xlsx_path)
    ws = wb[_pick_sheet_name(str(xlsx_path), sheet_name)]
    header = [str(c.value or "").strip() for c in ws[1]]
    col_index = {name: idx + 1 for idx, name in enumerate(header)}
    id_col = col_index["ID"]
    by_post_row: dict[str, int] = {}
    for r in range(2, ws.max_row + 1):
        pid = str(ws.cell(row=r, column=id_col).value or "").strip()
        if pid:
            by_post_row[pid] = r
    for post_id, col_name, new_url in updates:
        row = by_post_row.get(str(post_id))
        if row is None:
            continue
        ws.cell(row=row, column=col_index[col_name], value=new_url)
        logger.info("Media sync: Excel #%s %s → %s", post_id, col_name, new_url)
    wb.save(xlsx_path)


def sync_media(source: str, sheet_name: str) -> bool:
    """
    Синхронизирует медиа таблицы с бакетом.
    Возвращает True, если таблица была изменена (нужно перечитать посты).
    """
    if not auto_sync_enabled():
        return False

    bucket = bucket_name()
    cells = load_media_cells(source, sheet_name)
    desired: dict[str, str] = {_cell_key(c.post_id, c.col): c.url for c in cells}

    with _state_lock:
        state = _load_state()
        state_cells: dict = state.setdefault("cells", {})
        s3 = get_s3_client()
        sheet_updates: list[tuple[str, str, str]] = []

        # 1) Пустые ячейки — удалить из бакета
        for key, url in desired.items():
            if url:
                continue
            post_id, col = key.split(":", 1)
            delete_cell(s3, bucket, post_id, col)
            state_cells.pop(key, None)

        # 2) Синхронизация непустых ячеек
        for cell in cells:
            if not cell.url:
                continue
            key = _cell_key(cell.post_id, cell.col)
            prev = state_cells.get(key, {})
            bucket_b = bucket

            if is_yandex_url(cell.url, bucket_b):
                ykey = yandex_url_to_key(cell.url, bucket_b)
                if ykey:
                    if _recompress_bucket_video_if_needed(s3, bucket_b, cell.post_id, cell.col, ykey):
                        ykey = object_key(cell.post_id, cell.col, "mp4")
                        cell_url = yandex_public_url(ykey, bucket_b)
                        state_cells[key] = {
                            "source_url": cell.url,
                            "yandex_key": ykey,
                            "yandex_url": cell_url,
                        }
                        if cell.url != cell_url:
                            sheet_updates.append((cell.post_id, cell.col, cell_url))
                        continue
                    state_cells[key] = {
                        "source_url": cell.url,
                        "yandex_key": ykey,
                        "yandex_url": cell.url,
                    }
                continue

            if not _should_sync_url(cell.url, bucket_b):
                logger.warning(
                    "Media sync: #%s %s — неизвестный URL, пропуск: %s",
                    cell.post_id,
                    cell.col,
                    cell.url[:80],
                )
                continue

            if prev.get("source_url") == cell.url and prev.get("yandex_url"):
                if not is_yandex_url(cell.url, bucket_b):
                    sheet_updates.append((cell.post_id, cell.col, prev["yandex_url"]))
                continue

            old_key = prev.get("yandex_key")
            if old_key:
                delete_keys(s3, bucket, [old_key])
            delete_cell(s3, bucket, cell.post_id, cell.col)

            with tempfile.TemporaryDirectory(prefix="dekor_sync_") as tmp:
                ext = guess_ext(cell.url, cell.kind)
                local = Path(tmp) / f"{cell.post_id}_{cell.col}.{ext}"
                logger.info("Media sync: скачивание #%s %s …", cell.post_id, cell.col)
                download_media(cell.url, local)
                if cell.kind == "video":
                    local, ext = _prepare_local_media(local, cell.kind)
                else:
                    ext = guess_ext(cell.url, cell.kind, local_path=local)
                ykey = object_key(cell.post_id, cell.col, ext)
                upload_file(s3, bucket, local, ykey)
                public = yandex_public_url(ykey, bucket)
                size_mb = local.stat().st_size / (1024 * 1024)
                logger.info("Media sync: залито %.1f МБ → %s", size_mb, public)

            sheet_updates.append((cell.post_id, cell.col, public))
            state_cells[key] = {
                "source_url": cell.url,
                "yandex_key": ykey,
                "yandex_url": public,
            }

        # 3) Сироты в бакете (пост удалён из таблицы или ячейка пустая)
        for obj_key in list_objects(s3, bucket, "posts/"):
            parsed = _parse_bucket_cell_key(obj_key)
            if parsed is None:
                continue
            pid, col = parsed
            ck = _cell_key(pid, col)
            if ck not in desired or not desired[ck]:
                delete_keys(s3, bucket, [obj_key])
                state_cells.pop(ck, None)

        _save_state(state)

    if sheet_updates:
        update_sheet_cells(source, sheet_name, sheet_updates)
        return True
    return False


def sync_media_if_enabled(source: str, sheet_name: str) -> bool:
    try:
        return sync_media(source, sheet_name)
    except Exception:
        logger.exception("Media sync: ошибка синхронизации с бакетом")
        return False


def main() -> None:
    import logging

    from dotenv import load_dotenv

    load_dotenv()
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s [%(levelname)s] %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
    )
    source = os.getenv("POSTS_XLSX_PATH", "").strip()
    sheet = os.getenv("POSTS_SHEET_NAME", "Posts").strip()
    if sync_media(source, sheet):
        logger.info("Media sync: таблица обновлена.")
    else:
        logger.info("Media sync: изменений нет.")


if __name__ == "__main__":
    main()
